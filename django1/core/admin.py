import csv
from django.http import HttpResponse
from openpyxl import Workbook  # Para exportação em Excel
from django.contrib import admin
from .models import Pagamento
from django.db.models import Sum
from .models import Favorecido, Banco
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponseRedirect
from django.urls import reverse
from .models import Banco, HistoricoSaldo
from django.utils.safestring import mark_safe
from django.contrib import messages
from .models import  Recebimento, Transferencia,Pagador,CentroCusto
# Verifique se há outra chamada ao register para Banco


# Registro dos modelos no Django Admin
#admin.site.register(Socio)
#admin.site.register(Fornecedor)
#admin.site.register(Cliente)
#admin.site.register(Funcionario)
#admin.site.register(Favorecido)

#if not admin.site.is_registered(Banco):
    #admin.site.register(Banco)
#admin.site.register(Banco)

# Personalização do painel de administração
admin.site.site_header = "Administração do Sistema"
admin.site.site_title = "Administração VIP Air"
admin.site.index_title = "Bem-vindo ao Painel de Controle"




   

'''
class PagamentoAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'descricao', 'favorecido', 'nome','banco', 'data_prevista',
        'data_vencimento','data_pagamento', 'valor','categoria' ,'Responsavel','status'
    )
    list_filter = ('status', 'banco', 'data_vencimento','data_pagamento')
    search_fields = ('descricao', 'favorecido')

    # Adicionando as ações de exportação
    actions = ['exportar_para_csv', 'exportar_para_excel']

    def changelist_view(self, request, extra_context=None):
        queryset = self.get_queryset(request).distinct()  # Garantir que não há duplicação
        total_valor = queryset.aggregate(Sum('valor'))['valor__sum'] or 0
        extra_context = extra_context or {}
        extra_context['total_valor'] = total_valor
        return super().changelist_view(request, extra_context=extra_context)

    # Ação para exportar os dados para CSV
    @admin.action(description="Exportar dados para CSV")
    def exportar_para_csv(self, request, queryset):
        queryset = queryset.distinct()  # Garantir que não há duplicação no queryset

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="pagamentos.csv"'

        writer = csv.writer(response)
        writer.writerow(['ID', 'Descrição', 'Favorecido','Nome', 'Banco', 'Data Prevista', 'Data Vencimento', 'Data pagamento','Valor','Categoria' ,'Responsavel', 'Status'])

        for pagamento in queryset:
            writer.writerow([
                pagamento.id,
                pagamento.descricao,
                pagamento.favorecido,
                pagamento.nome,
                pagamento.banco.nome,  # Acessa o atributo 'nome' do modelo Banco
                pagamento.data_prevista,
                pagamento.data_vencimento,
                pagamento.data_pagamento,
                pagamento.valor,
                pagamento.categoria,
                pagamento.responsavel,
                pagamento.status,
            ])

        return response

    # Ação para exportar os dados para Excel
    @admin.action(description="Exportar dados para Excel")
    def exportar_para_excel(self, request, queryset):
        queryset = queryset.distinct()  # Garantir que não há duplicação no queryset

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pagamentos.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Pagamentos"

        # Cabeçalho do Excel
        ws.append(['ID', 'Descrição', 'Favorecido', 'Nome','Banco', 'Data Prevista', 'Data Vencimento','Data pagamento', 'Valor','Categoria' ,'Responsavel', 'Status'])

        # Dados do Excel
        for pagamento in queryset:
            ws.append([
                pagamento.id,
                pagamento.descricao,
                pagamento.nome,
                pagamento.favorecido,
                pagamento.banco.nome,  # Acessa o atributo 'nome' do modelo Banco
                pagamento.data_prevista,
                pagamento.data_vencimento,
                pagamento.data_pagamento,
                pagamento.valor,
                pagamento.categoria,
                pagamento.responsavel,
                pagamento.status,
            ])

        wb.save(response)
        return response

# Registro do modelo 'Pagamento' com as configurações de administração
admin.site.register(Pagamento, PagamentoAdmin)





'''

@admin.register(CentroCusto)
class CentroCustoAdmin(admin.ModelAdmin):
    list_display = ('nome','descricao')
    search_fields = ('nome','descricao')

@admin.register(Pagamento)
class PagamentoAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'descricao', 'favorecido', 'get_nome', 'banco', 'data_prevista',
        'data_vencimento', 'data_pagamento', 'valor','centro_custo', 'categoria', 'get_responsavel', 'status'
    )
    list_filter = ('status', 'banco', 'data_vencimento', 'data_pagamento','centro_custo')
    search_fields = ('descricao', 'favorecido')

    # Adicionando as ações de exportação
    actions = ['modificar_pagamento', 'exportar_para_csv', 'exportar_para_excel', 'remover_pagamentos_selecionados', 'marcar_como_pago']

    def changelist_view(self, request, extra_context=None):
        """
        Adiciona o total do valor na lista de pagamentos.
        """
        queryset = self.get_queryset(request).distinct()  # Garantir que não há duplicação
        total_valor = queryset.aggregate(Sum('valor'))['valor__sum'] or 0
        extra_context = extra_context or {}
        extra_context['total_valor'] = total_valor
        return super().changelist_view(request, extra_context=extra_context)

    # Ação para exportar os dados para CSV
    @admin.action(description="Exportar dados para CSV")
    def exportar_para_csv(self, request, queryset):
        queryset = queryset.distinct()  # Garantir que não há duplicação no queryset

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="pagamentos.csv"'

        writer = csv.writer(response)
        writer.writerow(['ID', 'Descrição', 'Favorecido', 'Nome', 'Banco', 'Data Prevista',
                         'Data Vencimento', 'Data Pagamento', 'Valor','Centro_custo', 'Categoria', 'Responsavel', 'Status'])

        for pagamento in queryset:
            writer.writerow([
                pagamento.id,
                pagamento.descricao,
                pagamento.favorecido,
                pagamento.favorecido.nome,  # Corrigido para acessar o nome do favorecido
                pagamento.banco.nome,
                pagamento.data_prevista,
                pagamento.data_vencimento,
                pagamento.data_pagamento,
                pagamento.valor,
                pagamento.centro_custo.nome,
                pagamento.categoria,
                pagamento.responsavel,
                pagamento.status,
            ])

        return response

    # Ação para exportar os dados para Excel
    @admin.action(description="Exportar dados para Excel")
    def exportar_para_excel(self, request, queryset):
        queryset = queryset.distinct()  # Garantir que não há duplicação no queryset

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pagamentos.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Pagamentos"

        # Cabeçalho do Excel
        ws.append(['ID', 'Descrição', 'Favorecido', 'Nome', 'Banco', 'Data Prevista',
                   'Data Vencimento', 'Data Pagamento', 'Valor','Centro_custo', 'Categoria', 'Responsavel', 'Status'])

        # Dados do Excel
        for pagamento in queryset:
            ws.append([
                pagamento.id,
                pagamento.descricao,
                pagamento.favorecido,
                pagamento.favorecido.nome,  # Corrigido para acessar o nome do favorecido
                pagamento.banco.nome,
                pagamento.data_prevista,
                pagamento.data_vencimento,
                pagamento.data_pagamento,
                pagamento.valor,
                pagamento.centro_custo.nome,
                pagamento.categoria,
                pagamento.responsavel,
                pagamento.status,
            ])

        wb.save(response)
        return response

    # Controle dinâmico dos campos exibidos no formulário
    def get_fields(self, request, obj=None):
        """
        Exibe campos adicionais apenas se o status for 'pago'.
        """
        fields = [
            'descricao', 'responsavel', 'favorecido', 'observacao',
            'data_prevista', 'data_vencimento', 'valor','centro_custo', 'data',
            'status'
        ]
        if obj and obj.status == 'pago':
            fields.extend(['data_pagamento', 'arquivo', 'banco'])  # Exibir campos de data_pagamento e arquivo
        return fields

    def get_readonly_fields(self, request, obj=None):
        """
        Torna alguns campos somente leitura com base no status.
        """
        readonly_fields = []
        # Tornar 'status' somente leitura apenas quando o status não for 'pago' nem 'pendente'
        if obj and obj.status not in ['pago', 'pendente']:
            readonly_fields.append('status')  # 'status' será somente leitura para outros valores além de 'pago' e 'pendente'
        return readonly_fields

    def save_model(self, request, obj, form, change):
        """
        Preenche automaticamente a data_pagamento quando o status for 'pago'.
        """
        if obj.status == 'pago' and not obj.data_pagamento:
            from datetime import date
            obj.data_pagamento = date.today()  # Preencher a data de pagamento automaticamente
        super().save_model(request, obj, form, change)

    @admin.action(description="Modificar pagamento selecionado")
    def modificar_pagamento(modeladmin, request, queryset):
        """
        Redireciona para a página de edição do pagamento selecionado.
        """
        if queryset.count() == 1:
            pagamento = queryset.first()
            url = reverse('admin:core_pagamento_change', args=[pagamento.pk])
            return HttpResponseRedirect(url)
        else:
            modeladmin.message_user(
                request,
                _("Selecione apenas um pagamento para modificar."),
                level='error'
            )

    @admin.action(description='Marcar pagamentos selecionados como pagos')
    def marcar_como_pago(self, request, queryset):
        for pagamento in queryset:
            if pagamento.status != 'pago':
                pagamento.status = 'pago'
                pagamento.save()

    # Método para exibir o nome do favorecido
    def get_nome(self, obj):
        return obj.favorecido.nome  # Corrigido para acessar o nome do favorecido
    get_nome.admin_order_field = 'favorecido__nome'  # Para ordenar por esse campo
    get_nome.short_description = 'Nome'

    # Método para exibir o responsável
    def get_responsavel(self, obj):
        return obj.responsavel
    get_responsavel.admin_order_field = 'responsavel'  # Para ordenar por esse campo
    get_responsavel.short_description = 'Responsável'
    def save_model(self, request, obj, form, change):
        if 'status' in form.changed_data:
            if obj.status == 'pago':
                messages.info(request, f"O pagamento de {obj.valor} foi marcado como Pago.")
            elif obj.status == 'pendente':
                messages.warning(request, f"O pagamento de {obj.valor} foi marcado como Pendente e o saldo foi restaurado.")
        super().save_model(request, obj, form, change)





@admin.register(Pagador)
class PagadorAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'nome', 'razao_social', 'cnpj', 'cpf', 'nome_fantasia',
        'telefone', 'endereco', 'bairro', 'cidade', 'estado', 
        'pais', 'banco', 'agencia', 'conta', 'tipodeconta',
        'conta_pix', 'categoria'
    )
    list_filter = ('categoria', 'estado', 'cidade', 'banco')
    search_fields = ('nome', 'razao_social', 'cnpj', 'cpf', 'telefone')
    
    actions = ['exportar_para_csv', 'exportar_para_excel', 'modificar_pagador']

    # Adicionando suporte ao CSV
    @admin.action(description="Exportar dados para CSV")
    def exportar_para_csv(self, request, queryset):
        queryset = queryset.distinct()
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Pagadores.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Nome', 'Razão Social', 'CNPJ', 'CPF', 'Nome Fantasia',
            'Telefone', 'Endereço', 'Bairro', 'Cidade', 'Estado', 
            'País', 'Banco', 'Agência', 'Conta', 'Tipo de Conta',
            'Conta PIX', 'Categoria'
        ])

        for pagador in queryset:
            writer.writerow([
                pagador.id,
                pagador.nome,
                pagador.razao_social,
                pagador.cnpj,
                pagador.cpf,
                pagador.nome_fantasia,
                pagador.telefone,
                pagador.endereco,
                pagador.bairro,
                pagador.cidade,
                pagador.estado,
                pagador.pais,
                pagador.banco,
                pagador.agencia,
                pagador.conta,
                pagador.tipodeconta,
                pagador.conta_pix,
                pagador.categoria,
            ])
        
        return response

    # Adicionando suporte ao Excel
    @admin.action(description="Exportar dados para Excel")
    def exportar_para_excel(self, request, queryset):
        queryset = queryset.distinct()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Pagadores.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Pagadores"

        # Cabeçalho do Excel
        ws.append([
            'ID', 'Nome', 'Razão Social', 'CNPJ', 'CPF', 'Nome Fantasia',
            'Telefone', 'Endereço', 'Bairro', 'Cidade', 'Estado', 
            'País', 'Banco', 'Agência', 'Conta', 'Tipo de Conta',
            'Conta PIX', 'Categoria'
        ])

        # Dados do Excel
        for pagador in queryset:
            ws.append([
                pagador.id,
                pagador.nome,
                pagador.razao_social,
                pagador.cnpj,
                pagador.cpf,
                pagador.nome_fantasia,
                pagador.telefone,
                pagador.endereco,
                pagador.bairro,
                pagador.cidade,
                pagador.estado,
                pagador.pais,
                pagador.banco,
                pagador.agencia,
                pagador.conta,
                pagador.tipodeconta,
                pagador.conta_pix,
                pagador.categoria,
            ])
        
        wb.save(response)
        return response

    # Permitir Modificação Direta
    @admin.action(description="Modificar pagador selecionado")
    def modificar_pagador(modeladmin, request, queryset):
        if queryset.count() == 1:
            pagador = queryset.first()
            url = reverse('admin:core_pagador_change', args=[pagador.pk])
            return HttpResponseRedirect(url)
        else:
            modeladmin.message_user(
                request, 
                "Selecione apenas um pagador para modificar.",
                level='error'
            )


    # Controle de Campos Exibidos
    def get_fields(self, request, obj=None):
        """
        Define quais campos aparecem no formulário.
        """
        fields = [
            'nome', 'razao_social', 'cnpj', 'cpf', 'nome_fantasia',
            'telefone', 'endereco', 'numero', 'complemtento', 'bairro',
            'cidade', 'estado', 'pais', 'banco', 'agencia', 'conta',
            'tipodeconta', 'conta_pix', 'categoria'
        ]
        return fields

    # Controle de Campos Somente Leitura
    def get_readonly_fields(self, request, obj=None):
        """
        Define quais campos são somente leitura.
        """
        readonly_fields = []
        if obj and obj.categoria == 'Socio':  # Exemplo de regra de negócio
            readonly_fields.append('categoria')
        return readonly_fields
           
           




@admin.register(Favorecido)
class FavorecidoAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'nome', 'razao_social', 'cnpj', 'cpf', 'nome_fantasia',
        'telefone', 'endereco', 'bairro', 'cidade', 'estado', 
        'pais', 'banco', 'agencia', 'conta', 'tipodeconta',
        'conta_pix', 'categoria'
    )
    list_filter = ('categoria', 'estado', 'cidade', 'banco')
    search_fields = ('nome', 'razao_social', 'cnpj', 'cpf', 'telefone')
    
    actions = ['exportar_para_csv', 'exportar_para_excel', 'modificar_favorecido']

    # Adicionando suporte ao CSV
    @admin.action(description="Exportar dados para CSV")
    def exportar_para_csv(self, request, queryset):
        queryset = queryset.distinct()
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="favorecidos.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Nome', 'Razão Social', 'CNPJ', 'CPF', 'Nome Fantasia',
            'Telefone', 'Endereço', 'Bairro', 'Cidade', 'Estado', 
            'País', 'Banco', 'Agência', 'Conta', 'Tipo de Conta',
            'Conta PIX', 'Categoria'
        ])

        for favorecido in queryset:
            writer.writerow([
                favorecido.id,
                favorecido.nome,
                favorecido.razao_social,
                favorecido.cnpj,
                favorecido.cpf,
                favorecido.nome_fantasia,
                favorecido.telefone,
                favorecido.endereco,
                favorecido.bairro,
                favorecido.cidade,
                favorecido.estado,
                favorecido.pais,
                favorecido.banco,
                favorecido.agencia,
                favorecido.conta,
                favorecido.tipodeconta,
                favorecido.conta_pix,
                favorecido.categoria,
            ])
        
        return response

    # Adicionando suporte ao Excel
    @admin.action(description="Exportar dados para Excel")
    def exportar_para_excel(self, request, queryset):
        queryset = queryset.distinct()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="favorecidos.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Favorecidos"

        # Cabeçalho do Excel
        ws.append([
            'ID', 'Nome', 'Razão Social', 'CNPJ', 'CPF', 'Nome Fantasia',
            'Telefone', 'Endereço', 'Bairro', 'Cidade', 'Estado', 
            'País', 'Banco', 'Agência', 'Conta', 'Tipo de Conta',
            'Conta PIX', 'Categoria'
        ])

        # Dados do Excel
        for favorecido in queryset:
            ws.append([
                favorecido.id,
                favorecido.nome,
                favorecido.razao_social,
                favorecido.cnpj,
                favorecido.cpf,
                favorecido.nome_fantasia,
                favorecido.telefone,
                favorecido.endereco,
                favorecido.bairro,
                favorecido.cidade,
                favorecido.estado,
                favorecido.pais,
                favorecido.banco,
                favorecido.agencia,
                favorecido.conta,
                favorecido.tipodeconta,
                favorecido.conta_pix,
                favorecido.categoria,
            ])
        
        wb.save(response)
        return response

    # Permitir Modificação Direta
    @admin.action(description="Modificar favorecido selecionado")
    def modificar_favorecido(modeladmin, request, queryset):
        if queryset.count() == 1:
            favorecido = queryset.first()
            url = reverse('admin:core_favorecido_change', args=[favorecido.pk])
            return HttpResponseRedirect(url)
        else:
            modeladmin.message_user(
                request, 
                "Selecione apenas um favorecido para modificar.",
                level='error'
            )


    # Controle de Campos Exibidos
    def get_fields(self, request, obj=None):
        """
        Define quais campos aparecem no formulário.
        """
        fields = [
            'nome', 'razao_social', 'cnpj', 'cpf', 'nome_fantasia',
            'telefone', 'endereco', 'numero', 'complemtento', 'bairro',
            'cidade', 'estado', 'pais', 'banco', 'agencia', 'conta',
            'tipodeconta', 'conta_pix', 'categoria'
        ]
        return fields

    # Controle de Campos Somente Leitura
    def get_readonly_fields(self, request, obj=None):
        """
        Define quais campos são somente leitura.
        """
        readonly_fields = []
        if obj and obj.categoria == 'Socio':  # Exemplo de regra de negócio
            readonly_fields.append('categoria')
        return readonly_fields
           
           


@admin.register(Banco)
class BancoAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'nome', 'agencia', 'conta','saldo', 'tipodeconta', 'conta_pix', 'categoria')
    list_filter = ('categoria',)
    search_fields = ('nome', 'agencia', 'conta', 'categoria')

    actions = ['exportar_para_csv', 'exportar_para_excel', 'modificar_banco']

    # Ação para exportar os dados para CSV
    @admin.action(description="Exportar dados para CSV")
    def exportar_para_csv(self, request, queryset):
        queryset = queryset.distinct()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="bancos.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Nome', 'Agência', 'Conta','Saldo', 'Tipo de Conta', 'Conta PIX', 'Categoria'
        ])

        for banco in queryset:
            writer.writerow([
                banco.id,
                banco.nome,
                banco.agencia,
                banco.conta,
                banco.saldo,
                banco.tipodeconta,
                banco.conta_pix,
                banco.categoria,
            ])

        return response

    # Ação para exportar os dados para Excel
    @admin.action(description="Exportar dados para Excel")
    def exportar_para_excel(self, request, queryset):
        queryset = queryset.distinct()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bancos.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Bancos"

        # Cabeçalho do Excel
        ws.append([
            'ID', 'Nome', 'Agência', 'Conta', 'Saldo','Tipo de Conta', 'Conta PIX', 'Categoria'
        ])

        # Dados do Excel
        for banco in queryset:
            ws.append([
                banco.id,
                banco.nome,
                banco.agencia,
                banco.conta,
                banco.saldo,
                banco.tipodeconta,
                banco.conta_pix,
                banco.categoria,
            ])

        wb.save(response)
        return response

    # Permitir Modificação Direta
    @admin.action(description="Modificar banco selecionado")
    def modificar_banco(self, request, queryset):
        if queryset.count() == 1:
            banco = queryset.first()
            url = reverse('admin:core_banco_change', args=[banco.pk])
            return HttpResponseRedirect(url)
        else:
            self.message_user(
                request,
                "Selecione apenas um banco para modificar.",
                level='error'
            )

    # Controle de Campos Exibidos
    def get_fields(self, request, obj=None):
        """
        Define quais campos aparecem no formulário.
        """
        fields = [
            'nome', 'agencia', 'conta','saldo', 'tipodeconta', 'conta_pix', 'categoria'
        ]
        return fields

    # Controle de Campos Somente Leitura
    def get_readonly_fields(self, request, obj=None):
        """
        Define quais campos são somente leitura.
        """
        readonly_fields = []
        if obj and obj.categoria == 'Especial':  # Exemplo de regra de negócio
            readonly_fields.append('categoria')
        return readonly_fields
          



@admin.register(Recebimento)
class RecebimentoAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'descricao', 'pagador', 'data_recebimento', 'valor','centro_custo',
        'metodo_pagamento', 'status'
    )
    list_filter = ('status', 'data_recebimento', 'metodo_pagamento','centro_custo')
    search_fields = ('descricao', 'pagador', 'metodo_pagamento')
    actions = ['exportar_para_csv', 'exportar_para_excel', 'marcar_como_recebido']

    @admin.action(description="Exportar dados para CSV")
    def exportar_para_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="recebimentos.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Descrição', 'Pagador', 'Data de Recebimento',
            'Valor Recebido','Centro_custo', 'Método de Pagamento', 'Status'
        ])
        for recebimento in queryset:
            writer.writerow([
                recebimento.id, recebimento.descricao, recebimento.pagador,
                recebimento.data_recebimento, recebimento.valor,recebimento.centro_custo,
                recebimento.metodo_pagamento, recebimento.status,
            ])
        return response

    @admin.action(description="Exportar dados para Excel")
    def exportar_para_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="recebimentos.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Recebimentos"
        ws.append([
            'ID', 'Descrição', 'Pagador', 'Data de Recebimento',
            'Valor Recebido', 'Centro_custo','Método de Pagamento', 'Status'
        ])
        for recebimento in queryset:
            ws.append([
                recebimento.id, recebimento.descricao, recebimento.pagador,
                recebimento.data_recebimento, recebimento.valor,recebimento.centro_custo,
                recebimento.metodo_pagamento, recebimento.status,
            ])
        wb.save(response)
        return response


@admin.register(Transferencia)
class TransferenciaAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'banco_origem', 'banco_destino', 'valor', 'data_transferencia', 'status'
    )
    list_filter = ('status', 'data_transferencia')
    search_fields = ('banco_origem__nome', 'banco_destino__nome')

    actions = ['exportar_para_csv', 'exportar_para_excel']

    @admin.action(description="Exportar Transferências para CSV")
    def exportar_para_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transferencias.csv"'

        writer = csv.writer(response)
        writer.writerow(['ID', 'banco_origem', 'banco_destino', 'Valor', 'Data Transferência', 'Status'])

        for transferencia in queryset:
            writer.writerow([
                transferencia.id,
                transferencia.banco_origem.nome,
                transferencia.banco_destino.nome,
                transferencia.valor,
                transferencia.data_transferencia,
                transferencia.status
            ])
        return response

    @admin.action(description="Exportar Transferências para Excel")
    def exportar_para_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="transferencias.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Transferências"

        ws.append(['ID', 'Banco_origem', 'Banco_destino', 'Valor', 'Data Transferência', 'Status'])

        for transferencia in queryset:
            ws.append([
                transferencia.id,
                transferencia.banco_origem.nome,
                transferencia.banco_destino.nome,
                transferencia.valor,
                transferencia.data_transferencia,
                transferencia.status
            ])
        wb.save(response)
        return response


@admin.register(HistoricoSaldo)
class HistoricoSaldoAdmin(admin.ModelAdmin):
    list_display = ('banco', 'saldo_anterior', 'saldo_novo', 'data_alteracao')
    list_filter = ('banco', 'data_alteracao')
    search_fields = ('banco__nome',)
    change_list_template = 'admin/historico_saldo_change_list.html'
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        queryset = self.get_queryset(request).order_by('data_alteracao')
        extra_context['historico_data'] = list(queryset.values('data_alteracao', 'saldo_novo'))
        return super().changelist_view(request, extra_context=extra_context)
    class Media:
        js = [
            'https://cdn.jsdelivr.net/npm/chart.js',
        ]
#admin.site.register(Banco, BancoAdmin)
#admin.site.register(HistoricoSaldo)